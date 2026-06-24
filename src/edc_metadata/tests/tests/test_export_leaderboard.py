from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

import time_machine
from clinicedc_tests.consents import consent_v1
from clinicedc_tests.helper import Helper
from clinicedc_tests.visit_schedules.visit_schedule_metadata.visit_schedule import (
    get_visit_schedule,
)
from django.contrib.auth.models import User
from django.test import TestCase, override_settings, tag
from django.test.client import RequestFactory

from edc_consent import site_consents
from edc_facility.import_holidays import import_holidays
from edc_lab.models.panel import Panel
from edc_metadata.views.export_leaderboard_view import ExportLeaderboardView
from edc_visit_schedule.site_visit_schedules import site_visit_schedules

utc_tz = ZoneInfo("UTC")


@tag("metadata")
@override_settings(SITE_ID=10)
@time_machine.travel(datetime(2019, 8, 11, 8, 00, tzinfo=utc_tz))
class TestExportLeaderboard(TestCase):
    @classmethod
    def setUpTestData(cls):
        import_holidays()

    def setUp(self):
        self.user = User.objects.create(username="erik")
        for name in ["one", "two", "three", "four", "five", "six", "seven", "eight", "nine"]:
            Panel.objects.create(name=name)
        site_consents.registry = {}
        site_consents.register(consent_v1)
        site_visit_schedules._registry = {}
        site_visit_schedules.loaded = False
        visit_schedule = get_visit_schedule(consent_v1)
        site_visit_schedules.register(visit_schedule)

        helper = Helper()
        self.subject_visit = helper.enroll_to_baseline(
            visit_schedule_name=visit_schedule.name, schedule_name="schedule"
        )
        self.appointment = self.subject_visit.appointment
        self.vsn = self.appointment.visit_schedule_name
        self.sn = self.appointment.schedule_name

    def test_export_returns_xlsx_workbook(self):
        view = ExportLeaderboardView()
        view.request = RequestFactory().get(
            f"/export/?schedule={self.vsn}::{self.sn}&site=10"
        )
        view.request.user = self.user
        view.allowed_site_ids = lambda: [10]
        response = view.get(view.request)

        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml.sheet", response["Content-Type"])
        self.assertIn("attachment;", response["Content-Disposition"])

        from openpyxl import load_workbook  # noqa: PLC0415

        ws = load_workbook(BytesIO(response.content)).active
        self.assertEqual(ws.cell(row=1, column=1).value, "CRF")
        self.assertEqual(ws.cell(row=1, column=ws.max_column).value, "Total")
        # one header row + at least one CRF row
        self.assertGreater(ws.max_row, 1)
